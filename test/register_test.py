import unittest
import os
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import Workbook, load_workbook
import subprocess
import time

# Test data
TEST_DATA = {
    "TC_REGISTER_01": {
        "description": "Verify register with empty username",
        "username": "",
        "email": "test@example.com",
        "password": "Pass@123",
        "confirm_password": "Pass@123",
        "expected": "error:Username không được để trống!"
    },
    "TC_REGISTER_02": {
        "description": "Verify register with empty email",
        "username": "Test User",
        "email": "",
        "password": "Pass@123",
        "confirm_password": "Pass@123",
        "expected": "error:Email không được để trống!"
    },
    "TC_REGISTER_03": {
        "description": "Verify register with empty password",
        "username": "Test User",
        "email": "test@example.com",
        "password": "",
        "confirm_password": "Pass@123",
        "expected": "error:Mật khẩu không được để trống!"
    },
    "TC_REGISTER_04": {
        "description": "Verify register with empty confirm password",
        "username": "Test User",
        "email": "test@example.com",
        "password": "Pass@123",
        "confirm_password": "",
        "expected": "error:Nhập lại mật khẩu không được để trống!"
    },
    "TC_REGISTER_05": {
        "description": "Verify register with invalid email format",
        "username": "Test User",
        "email": "invalid_email",
        "password": "Pass@123",
        "confirm_password": "Pass@123",
        "expected": "error:Email không hợp lệ!"
    },
    "TC_REGISTER_06": {
        "description": "Verify register with password less than 8 characters",
        "username": "Test User",
        "email": "test@example.com",
        "password": "Pass@12",
        "confirm_password": "Pass@12",
        "expected": "error:Mật khẩu ít nhất 8  ký tự, chứa 1 số và 1 chữ in hoa!"
    },
    "TC_REGISTER_07": {
        "description": "Verify register with password missing uppercase letter",
        "username": "Test User",
        "email": "test@example.com",
        "password": "pass@123",
        "confirm_password": "pass@123",
        "expected": "error:Mật khẩu ít nhất 8 ký tự, chứa 1 số và 1 chữ in hoa!"
    },
    "TC_REGISTER_08": {
        "description": "Verify register with password missing number",
        "username": "Test User",
        "email": "test@example.com",
        "password": "Pass@abc",
        "confirm_password": "Pass@abc",
        "expected": "error:Mật khẩu ít nhất 8 ký tự, chứa 1 số và 1 chữ in hoa!"
    },
    "TC_REGISTER_09": {
        "description": "Verify register with mismatched passwords",
        "username": "Test User",
        "email": "test@example.com",
        "password": "Pass@123",
        "confirm_password": "Pass@124",
        "expected": "error:Mật khẩu không khớp!"
    },
    "TC_REGISTER_10": {
        "description": "Verify register with existing email",
        "username": "Test User",
        "email": "existing@example.com",
        "password": "Pass@123",
        "confirm_password": "Pass@123",
        "expected": "error:User already exists!"
    },
    "TC_REGISTER_11": {
        "description": "Verify register with SQL injection input",
        "username": "' OR '1'='1",
        "email": "' OR '1'='1@example.com",
        "password": "Pass@123",
        "confirm_password": "Pass@123",
        "expected": "error:Email không hợp lệ!"
    },
    "TC_REGISTER_12": {
        "description": "Verify register with valid data and successful navigation to login screen",
        "username": "Levantuan",
        "email": "levantuan123@gmail.com",
        "password": "Levantuan123@",
        "confirm_password": "Levantuan123@",
        "expected": "login_screen"
    }
}

# Đường dẫn file Excel
output_dir = "output"
file_path = os.path.join(output_dir, "test_results_register.xlsx")
results = []

# Tạo thư mục output nếu chưa có
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Nếu file không tồn tại -> tạo mới
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Test Case", "Description", "Result", "Status", "Note"])
    wb.save(file_path)

class TestRegisterAppium(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Khởi tạo kết nối Appium"""
        print("🔗 Connecting to Appium...")
        
        # Kiểm tra thiết bị Android
        try:
            devices = subprocess.check_output("adb devices", shell=True).decode()
            if "b9fff218" not in devices:
                raise Exception("Emulator-5554 not found. Please start the emulator.")
        except subprocess.CalledProcessError:
            raise Exception("Error running adb. Ensure Android SDK is installed and adb is in PATH.")

        # Cấu hình Appium
        options = UiAutomator2Options()
        options.platform_name = "Android"
        options.device_name = "MyAndroidDevice"
        options.udid = "b9fff218"
        options.app_package = "com.example.flutter_shop"
        options.app_activity = "com.example.flutter_shop.MainActivity"
        options.automation_name = "UiAutomator2"
        options.no_reset = True  # Không reset ứng dụng để giữ trạng thái

        # Thử kết nối với retry
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                cls.driver = webdriver.Remote("http://localhost:4723", options=options)
                cls.wait = WebDriverWait(cls.driver, 60)
                print("✅ Connected successfully!")
                break
            except Exception as e:
                print(f"⚠️ Appium connection error (attempt {attempt + 1}/{max_attempts}): {str(e)}")
                if attempt == max_attempts - 1:
                    raise Exception("Failed to connect to Appium after multiple attempts.")
                time.sleep(2)

        # Điều hướng tới màn hình đăng ký
        cls.navigate_to_register_screen()

    @classmethod
    def navigate_to_register_screen(cls):
        """Điều hướng tới màn hình đăng ký"""
        try:
            # Click vào Create an account
            create_account_button = cls.wait.until(EC.element_to_be_clickable(
                (AppiumBy.XPATH, "//android.widget.Button[@content-desc='Create an account']")
            ))
            create_account_button.click()
            print("🖱️ Clicked Create an account")
            time.sleep(2)

            # Kiểm tra xem đã ở màn hình đăng ký chưa (giữ nguyên logic cũ nếu cần)
            try:
                login_button = cls.wait.until(EC.element_to_be_clickable(
                    (AppiumBy.XPATH, "//android.widget.Button[@text='Already have an account? Login']")
                ))
                login_button.click()
                print("🖱️ Clicked to navigate to register screen")
                time.sleep(2)
            except TimeoutException:
                print("⚠️ Could not find button to navigate to register screen")
        except TimeoutException as e:
            print(f"⚠️ Navigation error: {str(e)}")

    @classmethod
    def tearDownClass(cls):
        """Đóng kết nối Appium và lưu kết quả"""
        cls.driver.quit()
        print("🔴 Disconnected!")
        wb = load_workbook(file_path)
        ws = wb.active
        for result in results:
            ws.append(result)
        wb.save(file_path)
        print(f"📝 Results saved to {file_path}")

    def clear_form(self):
        """Xóa dữ liệu trong form đăng ký"""
        try:
            fields = [
                "//android.widget.EditText[@index='1']",  # Username
                "//android.widget.EditText[@index='2']",  # Email
                "//android.widget.EditText[@index='3']",  # Password
                "//android.widget.EditText[@index='4']"   # Confirm Password
            ]
            for xpath in fields:
                field = self.wait.until(EC.visibility_of_element_located((AppiumBy.XPATH, xpath)))
                field.click()
                field.clear()
            print("🧹 Form cleared")
        except TimeoutException as e:
            print(f"⚠️ Could not clear form: {str(e)}")

    def register(self, username, email, password, confirm_password, expected_error=None):
        """Hàm thực hiện đăng ký"""
        try:
            print(f"🔍 Registering with Username: {username} / Email: {email} / Password: {password} / Confirm Password: {confirm_password}")
            # Trường Username
            username_field = self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='1']")
            ))
            username_field.click()
            username_field.clear()
            username_field.send_keys(username)

            # Trường Email
            email_field = self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='2']")
            ))
            email_field.click()
            email_field.clear()
            email_field.send_keys(email)

            # Trường Password
            pass_field = self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='3']")
            ))
            pass_field.click()
            pass_field.clear()
            pass_field.send_keys(password)

            # Trường Confirm Password
            confirm_pass_field = self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='4']")
            ))
            confirm_pass_field.click()
            confirm_pass_field.clear()
            confirm_pass_field.send_keys(confirm_password)
            time.sleep(2)

            # Nhấn nút Register
            register_button = self.wait.until(EC.element_to_be_clickable(
                (AppiumBy.XPATH, "//android.widget.Button[@content-desc='Register']")
            ))
            register_button.click()
            print("🖱️ Clicked Register button!")

            # Chờ ngắn hơn để kiểm tra thông báo lỗi hoặc chuyển hướng
            time.sleep(3)

            # Nếu có lỗi mong đợi, kiểm tra thông báo lỗi ngay sau khi nhấn Register
            if expected_error and expected_error != "login_screen":
                error_msg = expected_error.split("error:")[1]
                if self.is_error_message_displayed(error_msg):
                    print("❌ Đăng ký không thành công!")
                    return None
                else:
                    print("❌ Đăng ký không thành công! Không tìm thấy thông báo lỗi mong đợi.")
                    return "Error message not displayed"

            # Kiểm tra nếu chuyển hướng thành công tới màn hình đăng nhập
            if self.is_login_screen():
                if expected_error == "login_screen":
                    print("✅ Đăng ký thành công! Chuyển hướng đến màn hình đăng nhập")
                    return None
                else:
                    print("❌ Đăng ký không thành công! Chuyển hướng sai đến màn hình đăng nhập")
                    return "Unexpected navigation to login screen"
            else:
                print("❌ Đăng ký không thành công! Không chuyển hướng đến màn hình đăng nhập")
                return "Failed to redirect to login screen after registration"

        except TimeoutException as e:
            print("❌ Đăng ký không thành công!")
            return f"Error: Element not found - {str(e)}"
        except Exception as e:
            print("❌ Đăng ký không thành công!")
            return f"Unknown error: {str(e)}"

    def is_login_screen(self):
        """Kiểm tra xem có đang ở màn hình đăng nhập"""
        try:
            self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='1']")  # Trường email của login
            ))
            print("")
            return True
        except TimeoutException:
            print("❓ Not on login screen")
            return False

    def is_error_message_displayed(self, expected_message):
        """Kiểm tra thông báo lỗi với thời gian chờ ngắn hơn"""
        try:
            # Sử dụng WebDriverWait với thời gian chờ 3 giây
            WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, f"//android.widget.TextView[@text='{expected_message}']")
            ))
            print(f"⚠️ Error message found: '{expected_message}'")
            return True
        except TimeoutException:
            print(f"❌ Error message not found: '{expected_message}'")
            return False

    def navigate_back_to_register(self):
        """Điều hướng trở lại màn hình đăng ký sau mỗi test"""
        if self.is_login_screen():
            try:
                login_button = self.wait.until(EC.element_to_be_clickable(
                    (AppiumBy.XPATH, "//android.widget.Button[@text='Already have an account? Login']")
                ))
                login_button.click()
                print("🖱️ Navigated back to register screen")
                time.sleep(2)
            except TimeoutException:
                print("")
                self.navigate_to_register_screen()

    def test_register_sequential(self):
        """Chạy tuần tự tất cả các test case cho đến khi đăng ký thành công"""
        for test_id in TEST_DATA.keys():
            data = TEST_DATA[test_id]
            print(f"\n📋 Running test {test_id}: {data['description']}")
            
            # Gọi hàm register với thông báo lỗi mong đợi
            error = self.register(
                data["username"],
                data["email"],
                data["password"],
                data["confirm_password"],
                data["expected"]
            )
            expected = data["expected"]
            
            if expected == "login_screen":
                # Trường hợp tích cực: Mong đợi đăng ký thành công và chuyển hướng
                if error is None and self.is_login_screen():
                    result = True
                    status = "Pass"
                    note = "Successfully registered and redirected to login screen"
                    print("✅ Test PASS: Registration successful")
                else:
                    result = False
                    status = "Fail"
                    note = error if error else "Failed to reach login screen"
                    print(f"❌ Test FAIL: {note}")
            else:
                # Trường hợp tiêu cực: Kết quả dựa trên kiểm tra trong hàm register
                error_msg = expected.split("error:")[1]
                result = error is None  # Nếu không có lỗi trả về từ register, tức là tìm thấy thông báo lỗi
                status = "Pass" if result else "Fail"
                note = f"Expected error shown: '{error_msg}'" if result else f"Error message not displayed: '{error_msg}'"
                print(f"{'✅ Test PASS' if result else '❌ Test FAIL'}: {note}")

            results.append([test_id, data["description"], result, status, note])
            print(f"📊 Recorded result for {test_id}: {status}")

            # Nếu đăng ký thành công (TC_REGISTER_12), dừng lại
            if expected == "login_screen" and status == "Pass":
                break

            # Xóa form và tiếp tục test case tiếp theo
            self.clear_form()
            # Điều hướng lại màn hình đăng ký nếu cần
            if not self.is_login_screen():
                self.navigate_to_register_screen()

if __name__ == "__main__":
    unittest.main()