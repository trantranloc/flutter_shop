import unittest
import os
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, NoSuchDriverException
from openpyxl import Workbook, load_workbook
import subprocess
import time

# Test data
TEST_DATA = {
    "TC_CHECKOUT_01": {
        "description": "Verify checkout fails with missing Last Name and First Name",
        "last_name": "",  # Missing
        "first_name": "",  # Missing
        "phone": "0999888666",
        "address": "33 Xô Viết Nghệ Tĩnh",
        "city": "Đà Nẵng",
        "district": "Cẩm Lệ",
        "zip_code": "868866",
        "note": "Giao vào giờ trưa",
        "expected": "order_fail",
        "expected_error": "Required field missing"
    },
    "TC_CHECKOUT_02": {
        "description": "Verify checkout fails with missing Phone",
        "last_name": "Trần",
        "first_name": "Sang",
        "phone": "",  # Missing
        "address": "33 Xô Viết Nghệ Tĩnh",
        "city": "Đà Nẵng",
        "district": "Cẩm Lệ",
        "zip_code": "868866",
        "note": "Giao vào giờ trưa",
        "expected": "order_fail",
        "expected_error": "Required field missing"
    },
    "TC_CHECKOUT_03": {
        "description": "Verify checkout fails with missing Address",
        "last_name": "Trần",
        "first_name": "Sang",
        "phone": "0999888666",
        "address": "",  # Missing
        "city": "Đà Nẵng",
        "district": "Cẩm Lệ",
        "zip_code": "868866",
        "note": "Giao vào giờ trưa",
        "expected": "order_fail",
        "expected_error": "Required field missing"
    },
    "TC_CHECKOUT_04": {
        "description": "Verify checkout fails with missing City",
        "last_name": "Trần",
        "first_name": "Sang",
        "phone": "0999888666",
        "address": "33 Xô Viết Nghệ Tĩnh",
        "city": "",  # Missing
        "district": "Cẩm Lệ",
        "zip_code": "868866",
        "note": "Giao vào giờ trưa",
        "expected": "order_fail",
        "expected_error": "Required field missing"
    },
    "TC_CHECKOUT_05": {
        "description": "Verify checkout fails with missing District",
        "last_name": "Trần",
        "first_name": "Sang",
        "phone": "0999888666",
        "address": "33 Xô Viết Nghệ Tĩnh",
        "city": "Đà Nẵng",
        "district": "",  # Missing
        "zip_code": "868866",
        "note": "Giao vào giờ trưa",
        "expected": "order_fail",
        "expected_error": "Required field missing"
    },
    "TC_CHECKOUT_06": {
        "description": "Verify checkout fails with missing Zip Code",
        "last_name": "Trần",
        "first_name": "Sang",
        "phone": "0999888666",
        "address": "33 Xô Viết Nghệ Tĩnh",
        "city": "Đà Nẵng",
        "district": "Cẩm Lệ",
        "zip_code": "",  # Missing
        "note": "Giao vào giờ trưa",
        "expected": "order_fail",
        "expected_error": "Required field missing"
    },
    "TC_CHECKOUT_07": {
        "description": "Verify adding products, modifying quantities, removing product, and completing checkout successfully",
        "last_name": "Trần",
        "first_name": "Sang",
        "phone": "0999888666",
        "address": "33 Xô Viết Nghệ Tĩnh",
        "city": "Đà Nẵng",
        "district": "Cẩm Lệ",
        "zip_code": "868866",
        "note": "Giao vào giờ trưa",
        "expected": "order_success",
        "expected_error": None
    }
}

# Đường dẫn file Excel
output_dir = "output"
file_path = os.path.join(output_dir, "test_results_checkout.xlsx")
results = []

# Tạo thư mục output nếu chưa có
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Nếu file không tồn tại -> tạo mới
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Test Case", "Description", "Result", "Status", "Note"])
    wb.save(file_path)

class TestCheckoutAppium(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Khởi tạo kết nối Appium"""
        print("🔗 Connecting to Appium...")
        
        # Kiểm tra thiết bị Android
        try:
            devices = subprocess.check_output("adb devices", shell=True).decode()
            if "192.168.154.102:5555" not in devices:
                raise Exception("Emulator-5554 not found. Please start the emulator.")
        except subprocess.CalledProcessError:
            raise Exception("Error running adb. Ensure Android SDK is installed and adb is in PATH.")

        # Cấu hình Appium
        options = UiAutomator2Options()
        options.platform_name = "Android"
        options.device_name = "MyAndroidDevice"
        options.udid = "192.168.154.102:5555"
        options.app_package = "com.example.flutter_shop"
        options.app_activity = "com.example.flutter_shop.MainActivity"
        options.automation_name = "UiAutomator2"
        options.no_reset = True
        options.new_command_timeout = 300

        # Thử kết nối với retry
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                cls.driver = webdriver.Remote("http://localhost:4723", options=options)
                cls.wait = WebDriverWait(cls.driver, 60)
                print("✅ Connected successfully!")
                print("⏳ Waiting for 55 seconds before starting test...")
                time.sleep(55)
                # Kiểm tra session còn hoạt động không
                if not cls.driver.session_id:
                    raise Exception("Session ID is invalid")
                cls.driver.current_activity
                print("✅ Session is active")
                print("🚀 Starting test execution...")
                break
            except (NoSuchDriverException, Exception) as e:
                print(f"⚠️ Appium connection error (attempt {attempt + 1}/{max_attempts}): {str(e)}")
                if attempt == max_attempts - 1:
                    raise Exception("Failed to connect to Appium after multiple attempts.")
                time.sleep(5)

    @classmethod
    def tearDownClass(cls):
        """Đóng kết nối Appium và lưu kết quả"""
        try:
            if hasattr(cls, 'driver') and cls.driver:
                cls.driver.quit()
                print("🔴 Disconnected!")
        except Exception as e:
            print(f"⚠️ Error during driver quit: {str(e)}")
        wb = load_workbook(file_path)
        ws = wb.active
        for result in results:
            ws.append(result)
        wb.save(file_path)
        print(f"📝 Results saved to {file_path}")

    def find_element_with_retry(self, by, value, retries=5, wait_time=15):
        """Tìm phần tử với cơ chế thử lại"""
        for attempt in range(retries):
            try:
                element = self.wait.until(EC.element_to_be_clickable((by, value)))
                print(f"🔍 Found element: {value}")
                is_displayed = element.is_displayed()
                is_enabled = element.is_enabled()
                print(f"ℹ️ Element state - Displayed: {is_displayed}, Enabled: {is_enabled}")
                return element
            except TimeoutException:
                print(f"⚠️ Attempt {attempt + 1}/{retries}: Could not find element {value}")
                time.sleep(2)
        raise NoSuchElementException(f"Could not find element {value} after {retries} attempts")

    def slow_swipe_up(self):
        """Thực hiện thao tác vuốt lên chậm để cuộn trang chính xác"""
        size = self.driver.get_window_size()
        start_x = size["width"] * 0.5
        start_y = size["height"] * 0.7
        end_y = size["height"] * 0.4
        self.driver.swipe(start_x, start_y, start_x, end_y, 1000)
        print("📜 Slow swiped up to scroll page")
        time.sleep(1.5)

    def scroll_to_top(self):
        """Cuộn trang checkout về đầu trang"""
        size = self.driver.get_window_size()
        start_x = size["width"] * 0.5
        start_y = size["height"] * 0.4
        end_y = size["height"] * 0.7
        self.driver.swipe(start_x, start_y, start_x, end_y, 1000)
        print("📜 Scrolled to top of checkout page")
        time.sleep(1.5)

    def scroll_to_element(self, xpath):
        """Cuộn trang đến phần tử được chỉ định bởi XPath sử dụng UiAutomator hoặc swipe"""
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                element = self.driver.find_element(AppiumBy.XPATH, xpath)
                if element.is_displayed():
                    print(f"📜 Element already visible: {xpath}")
                    return
                content_desc = element.get_attribute("content-desc") or element.get_attribute("text") or ""
                if content_desc:
                    self.driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,
                        f'new UiSelector().descriptionContains("{content_desc}")'
                    )
                    print(f"📜 Scrolled to element: {xpath} (content-desc: {content_desc})")
                    time.sleep(1.5)
                    return
                else:
                    print(f"⚠️ Element {xpath} has no content-desc or text, attempting slow swipe")
                    self.slow_swipe_up()
            except NoSuchElementException:
                print(f"⚠️ Attempt {attempt + 1}/{max_attempts}: Could not find element {xpath}, attempting slow swipe")
                self.slow_swipe_up()
            except ValueError as e:
                print(f"⚠️ Scroll error: {str(e)}, attempting slow swipe")
                self.slow_swipe_up()
        raise NoSuchElementException(f"Could not scroll to element {xpath} after {max_attempts} attempts")

    def return_to_checkout_page(self):
        """Quay lại trang checkout bằng cách nhấn nút Back hoặc điều hướng lại từ giỏ hàng"""
        try:
            # Thử nhấn Back để quay lại trang checkout
            self.driver.back()
            time.sleep(1)
            self.wait.until(EC.presence_of_element_located((
                AppiumBy.XPATH, "//android.widget.ScrollView//android.widget.EditText"
            )))
            print("✅ Returned to checkout page")
            self.scroll_to_top()  # Cuộn về đầu trang sau khi quay lại
        except TimeoutException:
            print("⚠️ Could not return to checkout page, navigating from cart")
            # Quay lại giỏ hàng
            self.driver.back()
            time.sleep(1)
            # Tìm và nhấn nút Checkout
            self.scroll_to_element("//android.widget.Button[@content-desc='Checkout']")
            checkout_button = self.find_element_with_retry(
                AppiumBy.XPATH,
                "//android.widget.Button[@content-desc='Checkout']",
                retries=5,
                wait_time=20
            )
            checkout_button.click()
            time.sleep(2)
            self.wait.until(EC.presence_of_element_located((
                AppiumBy.XPATH, "//android.widget.ScrollView//android.widget.EditText"
            )))
            print("✅ Returned to checkout page")
            self.scroll_to_top()  # Cuộn về đầu trang sau khi quay lại

    def test_checkout_flow(self):
        """Test cases for checkout with failure and success scenarios after cart actions"""
        print("\n📋 Starting cart actions for all test cases...")

        try:
            # Kiểm tra xem ứng dụng đã ở màn hình chính chưa
            try:
                self.wait.until(EC.presence_of_element_located((
                    AppiumBy.XPATH, "//android.widget.ImageView[contains(@content-desc, 'Arrangement') or contains(@content-desc, 'Bouquet')]"
                )))
                print("✅ App is on the main product screen")
            except TimeoutException:
                raise Exception("App did not load the main product screen")

            # Thêm sản phẩm 1, 2, 3, 4 vào giỏ hàng
            products = [
                ("White Rose Arrangement", "//android.widget.ImageView[contains(@content-desc, 'White Rose Arrangement')]/android.widget.Button"),
                ("Sunflower Bouquet", "//android.widget.ImageView[contains(@content-desc, 'Sunflower Bouquet')]/android.widget.Button"),
                ("White Lily Bouquet", "//android.widget.ImageView[contains(@content-desc, 'White Lily Bouquet')]/android.widget.Button"),
                ("Pink Tulip Bouquet", "//android.widget.ImageView[contains(@content-desc, 'Pink Tulip Bouquet')]/android.widget.Button")
            ]
            for product_name, product_xpath in products:
                button = self.find_element_with_retry(AppiumBy.XPATH, product_xpath)
                button.click()
                print(f"🛒 Added {product_name} to cart")
                time.sleep(2)

            # Vào giỏ hàng
            cart_button = self.find_element_with_retry(
                AppiumBy.XPATH,
                "//android.widget.FrameLayout[@resource-id='android:id/content']//android.widget.Button"
            )
            cart_button.click()
            print("🛍️ Navigated to cart")
            time.sleep(2)

            # Tăng số lượng sản phẩm 1 và 2 lên 3 (nhấn nút tăng 2 lần mỗi sản phẩm)
            for product, xpath in [
                ("White Rose Arrangement", "//android.widget.ImageView[contains(@content-desc, 'White Rose Arrangement')]/android.view.View[2]"),
                ("Sunflower Bouquet", "//android.widget.ImageView[contains(@content-desc, 'Sunflower Bouquet')]/android.view.View[2]")
            ]:
                for _ in range(2):
                    increase_button = self.find_element_with_retry(AppiumBy.XPATH, xpath)
                    increase_button.click()
                    print(f"➕ Increased quantity for {product}")
                    time.sleep(1)

            # Xóa sản phẩm thứ 4
            remove_button = self.find_element_with_retry(
                AppiumBy.XPATH,
                "//android.widget.ImageView[contains(@content-desc, 'Pink Tulip Bouquet')]/android.widget.Button"
            )
            remove_button.click()
            print("🗑️ Removed Pink Tulip Bouquet from cart")
            time.sleep(2)

            # Cuộn đến nút Checkout
            self.scroll_to_element("//android.widget.Button[@content-desc='Checkout']")
            checkout_button = self.find_element_with_retry(
                AppiumBy.XPATH,
                "//android.widget.Button[@content-desc='Checkout']",
                retries=5,
                wait_time=20
            )
            checkout_button.click()
            print("🛵 Proceeded to checkout")
            time.sleep(2)

            # Kiểm tra xem đã ở màn hình checkout chưa
            try:
                self.wait.until(EC.presence_of_element_located((
                    AppiumBy.XPATH, "//android.widget.ScrollView//android.widget.EditText"
                )))
                print("✅ Checkout screen loaded")
            except TimeoutException:
                raise Exception("Checkout screen did not load")

            # Thực hiện các test case trên trang checkout
            for test_case, data in TEST_DATA.items():
                print(f"\n📋 Running test {test_case}: {data['description']}")

                try:
                    # Nhập dữ liệu vào các trường, cuộn đến từng trường
                    fields = [
                        ("Last Name", "//android.widget.EditText[contains(@text, 'New') or @index='1']", data["last_name"]),
                        ("First Name", "//android.widget.EditText[contains(@text, 'User') or @index='2']", data["first_name"]),
                        ("Phone", "//android.widget.ScrollView/android.widget.EditText[4]", data["phone"]),
                        ("Address", "//android.widget.ScrollView/android.widget.EditText[5]", data["address"]),
                        ("City", "//android.widget.ScrollView/android.widget.EditText[6]", data["city"]),
                        ("District", "//android.widget.ScrollView/android.widget.EditText[7]", data["district"]),
                        ("Zip Code", "//android.widget.ScrollView/android.widget.EditText[8]", data["zip_code"]),
                        ("Note", "//android.widget.ScrollView/android.widget.EditText[9]", data["note"])
                    ]
                    for field_name, xpath, value in fields:
                        self.scroll_to_element(xpath)
                        field = self.find_element_with_retry(AppiumBy.XPATH, xpath, wait_time=20, retries=5)
                        field.click()
                        field.clear()
                        print(f"🧹 Cleared {field_name}")
                        if value:  # Chỉ nhập nếu giá trị không rỗng
                            field.send_keys(value)
                            entered_value = field.get_attribute("text")
                            if entered_value != value:
                                print(f"⚠️ Warning: Entered {field_name} value '{entered_value}' does not match expected '{value}'")
                            print(f"✍️ Entered {field_name}: {value}")
                        else:
                            print(f"✍️ Skipped {field_name}: Left empty")
                        time.sleep(1.5)

                    # Cuộn đến nút Đặt Hàng và nhấn chỉ một lần
                    self.scroll_to_element("//android.widget.Button[@content-desc='Đặt Hàng']")
                    place_order_button = self.find_element_with_retry(
                        AppiumBy.XPATH,
                        "//android.widget.Button[@content-desc='Đặt Hàng']",
                        retries=5,
                        wait_time=15
                    )
                    place_order_button.click()
                    print("📦 Placed order")
                    time.sleep(1)

                    # Kiểm tra kết quả
                    if data["expected"] == "order_success":
                        try:
                            success_message = self.wait.until(EC.visibility_of_element_located((
                                AppiumBy.XPATH,
                                "//android.widget.TextView[contains(@text, 'Order placed successfully')]"
                            )))
                            result = True
                            status = "Pass"
                            note = "Order placed successfully"
                            print("✅ Test PASS: Order placed successfully")
                        except TimeoutException:
                            result = False
                            status = "Fail"
                            note = "Failed to place order or success message not found"
                            print("❌ Test FAIL: Failed to place order")
                    else:
                        try:
                            error_message = self.wait.until(EC.visibility_of_element_located((
                                AppiumBy.XPATH,
                                f"//android.widget.TextView[contains(@text, '{data['expected_error']}')]"
                            )))
                            result = True
                            status = "Pass"
                            note = f"Expected failure: {data['expected_error']}"
                            print(f"✅ Test PASS: Expected failure - {data['expected_error']}")
                        except TimeoutException:
                            result = False
                            status = "Fail"
                            note = f"Expected error '{data['expected_error']}' not found"
                            print(f"❌ Test FAIL: Expected error not found")

                    # Quay lại trang checkout và cuộn về đầu trang cho test case tiếp theo (trừ test case thành công)
                    if test_case != "TC_CHECKOUT_07":
                        self.return_to_checkout_page()

                except Exception as e:
                    result = False
                    status = "Fail"
                    note = f"Error: {str(e)}"
                    print(f"❌ Test FAIL: {note}")
                    # Thử quay lại trang checkout nếu có lỗi (trừ test case thành công)
                    if test_case != "TC_CHECKOUT_07":
                        try:
                            self.return_to_checkout_page()
                        except Exception as nav_error:
                            note += f"; Failed to return to checkout: {str(nav_error)}"
                            print(f"⚠️ Failed to return to checkout: {str(nav_error)}")

                results.append([test_case, data["description"], result, status, note])
                print(f"📊 Recorded result for {test_case}: {status}")

        except Exception as e:
            # Ghi lại lỗi nếu quá trình chuẩn bị giỏ hàng thất bại
            result = False
            status = "Fail"
            note = f"Error in cart setup: {str(e)}"
            print(f"❌ Test FAIL: {note}")
            results.append(["SETUP", "Cart setup", result, status, note])

if __name__ == "__main__":
    unittest.main()