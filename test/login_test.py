import os
import time
import unittest
from openpyxl import Workbook, load_workbook
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# Test data
TEST_DATA = {
    "TC_LOGIN_01": {
        "description": "Verify login with a non-existent username",
        "email": "wronguser@example.com",
        "password": "Pass@123",
        "expected": "error:Invalid username or password"
    },
    "TC_LOGIN_02": {
        "description": "Verify login with a password missing an uppercase letter",
        "email": "user@example.com",
        "password": "pass@123",
        "expected": "error:Password must contain at least one uppercase letter!"
    },
    "TC_LOGIN_03": {
        "description": "Verify login with a password missing a special character",
        "email": "user@example.com",
        "password": "Pass123",
        "expected": "error:Password must contain at least one special character!"
    },
    "TC_LOGIN_04": {
        "description": "Verify login with a password missing a number",
        "email": "user@example.com",
        "password": "Pass@abc",
        "expected": "error:Password must contain at least one number!"
    },
    "TC_LOGIN_05": {
        "description": "Verify login with an empty username field",
        "email": "",
        "password": "Pass@123",
        "expected": "error:Email cannot be empty!"
    },
    "TC_LOGIN_06": {
        "description": "Verify login with an empty password field",
        "email": "user@example.com",
        "password": "",
        "expected": "error:Password cannot be empty!"
    },
    "TC_LOGIN_07": {
        "description": "Verify login with both username and password fields empty",
        "email": "",
        "password": "",
        "expected": "error:Email cannot be empty!"
    },
    "TC_LOGIN_08": {
        "description": "Verify login with special characters in the username",
        "email": "user@#$%@example.com",
        "password": "Pass@123",
        "expected": "error:Invalid email format!"
    },
    "TC_LOGIN_09": {
        "description": "Verify login with a username exceeding the maximum length",
        "email": "abcdefghijklmnopqrstuvwxyzabcdefghijklmnopqrstuvwxyzabcdefghijklmnopqrstuvwxyz@example.com",
        "password": "Pass@123",
        "expected": "error:Email is too long!"
    },
    "TC_LOGIN_10": {
        "description": "Verify login with a password exceeding the maximum length",
        "email": "user@example.com",
        "password": "Pass@123456789012345678901234567890123456789012345678901234567890123456789012345678901234567890123456789012345678901234567890123456789",
        "expected": "error:Password is too long!"
    },
    "TC_LOGIN_11": {
        "description": "Verify login with email starting with a number",
        "email": "123user@example.com",
        "password": "Pass@123",
        "expected": "error:Email cannot start with a number!"
    },
    "TC_LOGIN_12": {
        "description": "Verify login with valid username and password",
        "email": "user@example.com",
        "password": "Pass@123",
        "expected": "Success",
    },
}

# Đường dẫn file Excel
output_dir = "output"
file_path = os.path.join(output_dir, "test_results.xlsx")

# Tạo thư mục output nếu chưa có
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Nếu file không tồn tại -> tạo mới
is_new_file = not os.path.exists(file_path)
if is_new_file:
    wb = Workbook()
    ws = wb.active
    ws.append(["Test Case", "Description", "Username", "Password", "Expected Result"])
    wb.save(file_path)
    print("🆕 File mới đã được tạo, bắt đầu từ Test 1")

class TestLoginAppium(unittest.TestCase):
    results = {}  # Store results for the current test run

    @classmethod
    def setUpClass(cls):
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        # Đếm đúng số cột Test
        test_columns = [h for h in headers if h and str(h).strip().startswith("Test ")]
        # Nếu là file mới, bắt đầu từ Test 1, nếu không thì tăng số test
        cls.test_run_number = 1 if is_new_file else len(test_columns)
        cls.test_column_name = f"Test {cls.test_run_number}"

        if cls.test_column_name not in headers:
            ws.cell(row=1, column=len(headers) + 1).value = cls.test_column_name
            wb.save(file_path)

        print(f"🚀 Starting test run: {cls.test_column_name}")

    def setUp(self):
        options = UiAutomator2Options()
        options.platform_name = "Android"
        options.device_name = "MyAndroidDevice"
        options.udid = "b9fff218"
        options.app_package = "com.example.flutter_shop"
        options.app_activity = "com.example.flutter_shop.MainActivity"
        options.automation_name = "UiAutomator2"
        options.no_reset = True

        self.driver = webdriver.Remote("http://localhost:4723", options=options)
        self.wait = WebDriverWait(self.driver, 5)
        print(f"🔗 Driver created for test {self._testMethodName}")

    def tearDown(self):
        if hasattr(self, 'driver'):
            self.driver.quit()
            print(f"🔴 Driver closed for test {self._testMethodName}")

    @classmethod
    def tearDownClass(cls):
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        test_column_index = headers.index(cls.test_column_name) + 1

        # Initialize rows if empty
        if ws.max_row == 1:
            for test_id in TEST_DATA:
                data = TEST_DATA[test_id]
                ws.append([test_id, data["description"], data["email"], data["password"], data["expected"]])

        # Update test results for the current run
        for row_idx, test_id in enumerate(TEST_DATA, start=2):
            ws.cell(row=row_idx, column=test_column_index).value = cls.results.get(test_id, "Fail")

        wb.save(file_path)
        print(f"📝 Results saved to {file_path} for {cls.test_column_name}")

    def login(self, email, password):
        try:
            email_field = self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='1']")))
            email_field.click()
            email_field.clear()
            email_field.send_keys(email)

            pass_field = self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.widget.EditText[@index='2']")))
            pass_field.click()
            pass_field.clear()
            pass_field.send_keys(password)

            login_button = self.wait.until(EC.element_to_be_clickable(
                (AppiumBy.XPATH, "(//android.widget.Button)[1]")))
            login_button.click()
        except TimeoutException as e:
            return f"Error: Element not found - {str(e)}"
        except Exception as e:
            return f"Unknown error: {str(e)}"

    def is_home_screen(self):
        try:
            self.wait.until(EC.visibility_of_element_located(
                (AppiumBy.XPATH, "//android.view.View[@content-desc=\"LIRIS'FLORA\"]")))
            return True
        except TimeoutException:
            return False

    def run_test(self, test_id):
        start_time = time.time()
        data = TEST_DATA[test_id]
        print(f"\n📋 Running test {test_id}: {data['description']} ({self.test_column_name})")

        error = self.login(data["email"], data["password"])
        expected = data["expected"]

        if expected == "Success":
            if error is None and self.is_home_screen():
                status = "Pass"
                note = "Successfully logged in and redirected to home screen"
            else:
                status = "Fail"
                note = error if error else "Failed to reach home screen"
        else:
            if error and error == expected:
                status = "Pass"
                note = f"Expected error message received: {error}"
            else:
                status = "Fail"
                note = f"Expected error '{expected}' but got '{error}'"

        print(f"Status: {status}")
        self.results[test_id] = status
        # if status == "Fail":
        #     self.fail(note)

# Dynamically create test methods
def create_test_method(test_id):
    def test_method(self):
        self.run_test(test_id)
    test_method.__name__ = f"test_{test_id}"
    return test_method

# Add test methods to the class
for test_id in TEST_DATA:
    setattr(TestLoginAppium, f"test_{test_id}", create_test_method(test_id))

if __name__ == "__main__":
    unittest.main()